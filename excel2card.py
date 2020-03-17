import openpyxl
from PIL import Image,ImageDraw,ImageFont
import textwrap
import os


font = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W0.ttc', 30)
font_big = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W0.ttc', 60)
font_small = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W0.ttc', 20)


# 曜日の値を参照して色を生成する関数
def setColor(day):
    if day == None:
        return (255, 255, 255)
    elif day == "月":
        return (255, 248, 181)
    elif day == "火":
        return (235, 191, 175)
    elif day == "水":
        return (136, 190, 210)
    elif day == "木":
        return (186, 217, 158)
    else :
        return (144, 144, 144)

# エクセルを読み込んでカード画像を生成する関数
def drawCard():
    #エクセル読み込み（ファイル名・シート名はハードコード）
    workbook = openpyxl.load_workbook('cardsheet.xlsx')
    sheet = workbook["Sheet1"]

    for i in range(2,99):
        card_num = "card_"+str(i)
        # print(card_num)

        id = sheet.cell(row=i, column=1).value
        title = sheet.cell(row=i, column=2).value
        cost =sheet.cell(row=i, column=5).value
        description = sheet.cell(row=i, column=3).value
        flavor =sheet.cell(row=i, column=4).value
        dayOfWeek = flavor =sheet.cell(row=i, column=9).value

        color = setColor(dayOfWeek)

        print(id+","+title+","+dayOfWeek+","+str(color))

        im = Image.new("RGB",(708,1033),color)
        draw = ImageDraw.Draw(im)

        # draw.rectangle((2, 2, 706, 1030) ,outline=(0,0,0)) #Frame
        draw.rectangle((10, 10, 100, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Number
        draw.rectangle((110, 10, 610, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Title
        draw.rectangle((620, 10, 698, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Cost
        draw.rectangle((40, 110, 668, 500), fill=(255, 255, 255) ,outline=(0,0,0)) #illust
        draw.multiline_text((280, 260), "illust", fill=(0, 0, 0), font=font_big)
        draw.rectangle((40, 520, 668, 900), fill=(255, 255, 255) ,outline=(0,0,0))#description
        draw.rectangle((40, 920, 668, 1000), fill=(255, 255, 255) ,outline=(0,0,0))#flavor text


        draw.multiline_text((30, 30), id, fill=(0, 0, 0), font=font)

        #カード名の印字（折返し有り）
        wrap_list = textwrap.wrap(title, 16)

        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*40+10  # y座標をline_counterに応じて下げる
            draw.multiline_text((120, y+5),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        #カード効果の印字（折返し有り）
        wrap_list = textwrap.wrap(description, 20)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*70+80  # y座標をline_counterに応じて下げる
            draw.multiline_text((50, y+550),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        draw.multiline_text((640, 30), str(cost), fill=(0, 0, 0), font=font)

        #フレーバーテキストの印字（折返し有り）
        wrap_list = textwrap.wrap(description, 30)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*35+80  # y座標をline_counterに応じて下げる
            draw.multiline_text((50, y+850),line, fill=(0,0,0), font=font_small)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        #確認用
        im.show()

        #でかすぎるのでリサイズ
        im_resize = im.resize((int(im.width*0.48), int(im.height*0.48)),Image.LANCZOS)
        im_resize.save('card/'+id+'.jpg', quality=95)

if __name__ == '__main__':
    drawCard()
