import openpyxl
from PIL import Image,ImageDraw,ImageFont
import textwrap
import os

#実行時設定 Trueだと確認のみ。Falseだとカードを生成する。
flag_view=True

#カード枚数
num = 36
#フォント設定
font = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W3.ttc', 30)
font_big = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W3.ttc', 35)
font_small = ImageFont.truetype('/Library/Fonts/ヒラギノ角ゴシック W4.ttc', 20)


# 曜日の値を参照して色を生成する関数
def setColor(day):
    if day == None: #何も入ってなければまっしろ *入力漏れ発見用
        return (255, 255, 255)
    elif day == "1月":
        return (255, 242, 204)
    elif day == "2火":
        return (217, 225, 242)
    elif day == "3水":
        return (226, 239, 218)
    elif day == "4木":
        return (252, 228, 214)
    else : #わけのわからない値が入ってたらグレー *入力ミス発見用
        return (144, 144, 144)

# エクセルを読み込んで表面のカード画像を生成する関数
def generateFrontCard():
    print("generating card(front)...")

    #エクセル読み込み（ファイル名・シート名はハードコード）
    workbook = openpyxl.load_workbook('cardsheet.xlsx')
    sheet = workbook["sheet1"]

    # print("workbook:"+workbook.name)

    for i in range(2,num):
        id = sheet.cell(row=i, column=1).value
        title = sheet.cell(row=i, column=2).value
        cost = sheet.cell(row=i, column=4).value
        description = sheet.cell(row=i, column=3).value
        # flavor = sheet.cell(row=i, column=4).value
        dayOfWeek =sheet.cell(row=i, column=8).value

        color = setColor(dayOfWeek)

        #デバッグ用
        # print(id+","+title+","+description+","+dayOfWeek+","+str(color))

        im = Image.new("RGB",(708,1033),color)
        draw = ImageDraw.Draw(im)

        # draw.rectangle((2, 2, 706, 1030) ,outline=(0,0,0)) #Frame
        draw.rectangle((10, 10, 100, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Number
        draw.rectangle((110, 10, 610, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Title
        draw.rectangle((620, 10, 698, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Cost
        draw.rectangle((40, 110, 668, 500), fill=(255, 255, 255) ,outline=(0,0,0)) #illust
        draw.multiline_text((280, 260), "illust", fill=(0, 0, 0), font=font_big)
        draw.rectangle((40, 520, 668, 850), fill=(255, 255, 255) ,outline=(0,0,0))#description
        # draw.rectangle((40, 870, 668, 1015), fill=(255, 255, 255) ,outline=(0,0,0))#flavor text

        # im.show()

        #カードIDの印字
        draw.multiline_text((25, 30), id, fill=(0, 0, 0), font=font_big)

        #カード名の印字（折返し有り）
        wrap_list = textwrap.wrap(title, 16)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*40+10  # y座標をline_counterに応じて下げる
            draw.multiline_text((120, y+5),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        #カード効果の印字（折返し有り）
        wrap_list = textwrap.wrap(description, 20)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*40+80  # y座標をline_counterに応じて下げる
            draw.multiline_text((50, y+450),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        draw.multiline_text((645, 30), str(cost), fill=(0, 0, 0), font=font_big)

        #フレーバーテキストの印字（折返し有り）
        # wrap_list = textwrap.wrap(flavor, 30)
        # line_counter = 0  # 行数のカウンター
        # for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
        #     y = line_counter*35+80  # y座標をline_counterに応じて下げる
        #     draw.multiline_text((50, y+800),line, fill=(0,0,0), font=font_small)  # 1行分の文字列を画像に描画
        #     line_counter = line_counter +1  # 行数のカウンターに1

        if flag_view==True:
            # 確認用
            im.show()
            # print("flag_view is "+str(flag_view))

        else :
            #でかすぎるのでリサイズ　*LANCZOSなら劣化少なく縮小できるらしい
            im_resize = im.resize((int(im.width*0.48), int(im.height*0.48)),Image.LANCZOS)
            im_resize.save('front/'+id+'.jpg', quality=95)
            # print("flag_view is "+str(flag_view))
    print("done!")

#エクセルを読み込んで裏面を生成する関数
def generateRearCard():
    print("generating card(rear)...")

    #エクセル読み込み（ファイル名・シート名はハードコード）
    workbook = openpyxl.load_workbook('cardsheet.xlsx')
    sheet = workbook["sheet1"]

    for i in range(2,num):
        id = sheet.cell(row=i, column=1).value      #ID
        title = sheet.cell(row=i, column=2).value   #Title
        point = sheet.cell(row=i, column=5).value   #Point
        result = sheet.cell(row=i, column=6).value  #Result
        flavor = sheet.cell(row=i, column=7).value  #Flavor
        dayOfWeek =sheet.cell(row=i, column=8).value#DayOfWeek

        color = setColor(dayOfWeek)

        im = Image.new("RGB",(708,1033),color)
        draw = ImageDraw.Draw(im)

        draw.rectangle((10, 10, 100, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Number
        draw.rectangle((110, 10, 610, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Title
        draw.rectangle((620, 10, 698, 90), fill=(255, 255, 255) ,outline=(0,0,0)) #Point
        draw.rectangle((40, 110, 668, 500), fill=(255, 255, 255) ,outline=(0,0,0)) #Illust
        draw.multiline_text((280, 260), "illust", fill=(0, 0, 0), font=font_big)
        draw.rectangle((40, 520, 668, 800), fill=(255, 255, 255) ,outline=(0,0,0))#Result
        draw.rectangle((40, 820, 668, 1015), fill=(255, 255, 255) ,outline=(0,0,0))#flavor text


        draw.multiline_text((25, 30), id, fill=(0, 0, 0), font=font_big)

        #カード名の印字（折返し有り）
        wrap_list = textwrap.wrap(title, 16)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*40+10  # y座標をline_counterに応じて下げる
            draw.multiline_text((120, y+5),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        #結果の印字（折返し有り）
        wrap_list = textwrap.wrap(result, 20)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*40+80  # y座標をline_counterに応じて下げる
            draw.multiline_text((50, y+450),line, fill=(0,0,0), font=font)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        draw.multiline_text((625, 30), str(point), fill=(0, 0, 0), font=font_big)

        #フレーバーテキストの印字（折返し有り）
        wrap_list = textwrap.wrap(flavor, 30)
        line_counter = 0  # 行数のカウンター
        for line in wrap_list:  # wrap_listから1行づつ取り出しlineに代入
            y = line_counter*35+80  # y座標をline_counterに応じて下げる
            draw.multiline_text((50, y+750),line, fill=(0,0,0), font=font_small)  # 1行分の文字列を画像に描画
            line_counter = line_counter +1  # 行数のカウンターに1

        if flag_view==True:
            # 確認用
            # print("flag_view is "+str(flag_view))
            im.show()
        else :
            #でかすぎるのでリサイズ　*LANCZOSなら劣化少なく縮小できるらしい
            im_resize = im.resize((int(im.width*0.48), int(im.height*0.48)),Image.LANCZOS)
            im_resize.save('front/'+id+'.jpg', quality=95)
            # 確認用
            # print("flag_view is "+str(flag_view))

    print("done!")


if __name__ == '__main__':
    generateFrontCard()
    generateRearCard()
